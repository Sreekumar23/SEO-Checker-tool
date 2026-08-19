import concurrent.futures
import csv
import re
import sys
import unicodedata
from html.parser import HTMLParser
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from urllib.parse import urljoin, quote, urlparse, urlunparse
from pathlib import Path

# ADAudit Plus footer tabs
ADAUDITPLUS_ELEMENTS = [
    "active directory",
    "file server",
    "windows server",
    "workstation",
    "compliance",
    "related products",
]

# ADManager Plus footer tabs (all possible tabs across page types)
ADMANAGER_ELEMENTS = [
    "highlights",
    "ad management",
    "active directory reports",
    "exchange management",
    "popular products",
    "related products",
]

# Core tabs present on every ADManager Plus page — used for "Standard" check
ADMANAGER_CORE_ELEMENTS = [
    "highlights",
    "ad management",
    "active directory reports",
]

# ADSelfService Plus footer tabs (5 tabs: EN uses "Enterprise SSO" and "Self-service & security")
ADSELFSERVICE_ELEMENTS = [
    "password management",
    "adaptive mfa",
    "corporate sso",
    "self-service security",
    "related products",
]

# Core tabs for ADSelfService Plus "Standard" check
ADSELFSERVICE_CORE_ELEMENTS = [
    "password management",
    "adaptive mfa",
    "corporate sso",
]

# EventLog Analyzer footer tabs
EVENTLOG_ELEMENTS = [
    "log management",
    "it compliance",
    "log analyzer",
    "quick links",
    "related products",
]

# Only "related products" is required — other tabs vary by language/page type
EVENTLOG_CORE_ELEMENTS = [
    "related products",
]

# Log360 footer tabs — single tab only
LOG360_ELEMENTS = [
    "related products",
]

LOG360_CORE_ELEMENTS = [
    "related products",
]

# DataSecurity Plus footer tabs (5 fixed tabs across all pages/languages)
DATASECURITY_ELEMENTS = [
    "solutions",
    "regulatory compliance",
    "resources",
    "quick links",
    "related products",
]

DATASECURITY_CORE_ELEMENTS = [
    "solutions",
    "related products",
]

# Cloud SIEM footer tabs — no distinct tab structure; related products only
CLOUDSIEM_ELEMENTS = ["related products"]
CLOUDSIEM_CORE_ELEMENTS = ["related products"]

# Cloud Security Plus footer tabs (2 tabs: Highlights + Related Products)
CLOUDSECURITY_ELEMENTS = [
    "highlights",
    "related products",
]

CLOUDSECURITY_CORE_ELEMENTS = [
    "related products",
]

# Exchange Reporter Plus footer tabs — single tab only
EXCHANGEREPORTER_ELEMENTS = [
    "related products",
]

EXCHANGEREPORTER_CORE_ELEMENTS = [
    "related products",
]

# M365 Manager Plus footer tabs — single tab only
M365MANAGER_ELEMENTS = [
    "related products",
]

M365MANAGER_CORE_ELEMENTS = [
    "related products",
]

# RecoveryManager Plus footer tabs — single tab only
RECOVERYMANAGER_ELEMENTS = [
    "related products",
]

RECOVERYMANAGER_CORE_ELEMENTS = [
    "related products",
]

# SharePoint Manager Plus footer tabs — single tab only
SHAREPOINTMANAGER_ELEMENTS = [
    "related products",
]

SHAREPOINTMANAGER_CORE_ELEMENTS = [
    "related products",
]

# Shared element list for auto-detected single-tab products
SINGLE_TAB_ELEMENTS = ["related products"]

# Maps URL substrings → (display_name, all_footer_elements, standard_elements).
# First match wins. Add a new row here to support a new product — no other code changes needed.
PRODUCT_URL_MAP = [
    ('/self-service-password/',              'ADSelfService Plus',      ADSELFSERVICE_ELEMENTS,    ADSELFSERVICE_CORE_ELEMENTS),
    ('/adselfservice-plus/',                 'ADSelfService Plus',      ADSELFSERVICE_ELEMENTS,    ADSELFSERVICE_CORE_ELEMENTS),
    ('/ad-manager/',                         'ADManager Plus',          ADMANAGER_ELEMENTS,        ADMANAGER_CORE_ELEMENTS),
    ('/active-directory-manager-msp/',      'ADManager Plus MSP',      ADMANAGER_ELEMENTS,        ADMANAGER_CORE_ELEMENTS),
    ('/active-directory-audit/',             'ADAudit Plus',            ADAUDITPLUS_ELEMENTS,      ADAUDITPLUS_ELEMENTS),
    ('/adaudit-plus/',                       'ADAudit Plus',            ADAUDITPLUS_ELEMENTS,      ADAUDITPLUS_ELEMENTS),
    ('/eventlog/',                           'EventLog Analyzer',       EVENTLOG_ELEMENTS,         EVENTLOG_CORE_ELEMENTS),
    ('/log-management/',                     'Log360',                  LOG360_ELEMENTS,           LOG360_CORE_ELEMENTS),
    ('/log-management-mssp/',               'Log360',                  LOG360_ELEMENTS,           LOG360_CORE_ELEMENTS),
    ('/log360/',                             'Log360',                  LOG360_ELEMENTS,           LOG360_CORE_ELEMENTS),
    ('/data-security/',                      'DataSecurity Plus',       DATASECURITY_ELEMENTS,     DATASECURITY_CORE_ELEMENTS),
    ('/exchange-reports/',                   'Exchange Reporter Plus',  EXCHANGEREPORTER_ELEMENTS, EXCHANGEREPORTER_CORE_ELEMENTS),
    ('/microsoft-365-management-reporting/', 'M365 Manager Plus',       M365MANAGER_ELEMENTS,      M365MANAGER_CORE_ELEMENTS),
    ('/office365-management-reporting/',    'M365 Manager Plus',       M365MANAGER_ELEMENTS,      M365MANAGER_CORE_ELEMENTS),
    ('/m365-manager/',                       'M365 Manager Plus',       M365MANAGER_ELEMENTS,      M365MANAGER_CORE_ELEMENTS),
    ('/ad-recovery-manager/',                'RecoveryManager Plus',    RECOVERYMANAGER_ELEMENTS,  RECOVERYMANAGER_CORE_ELEMENTS),
    ('/sharepoint-management-reporting/',    'SharePoint Manager Plus', SHAREPOINTMANAGER_ELEMENTS, SHAREPOINTMANAGER_CORE_ELEMENTS),
    ('/microsoft-365-security-protection/',  'M365 Security Plus',      SINGLE_TAB_ELEMENTS,        SINGLE_TAB_ELEMENTS),
    ('/cloud-security/',                     'Cloud Security Plus',     CLOUDSECURITY_ELEMENTS,     CLOUDSECURITY_CORE_ELEMENTS),
    ('/cloud-siem/',                         'Cloud SIEM',              CLOUDSIEM_ELEMENTS,          CLOUDSIEM_CORE_ELEMENTS),
]

# Maps special multi-segment regional codes that are NOT ISO 639/3166 codes.
# Standard 2-letter codes (de, fr, dk…) are handled automatically by
# LANG_CODE_MAP + COUNTRY_LANG_MAP inside detect_language().
URL_LANG_MAP = [
    ('/pt-br/', 'Portuguese'),  # Brazilian Portuguese region code
    ('/latam/', 'Spanish'),     # Latin America region code
]

# ISO 639-1 language codes → display name
LANG_CODE_MAP = {
    'de': 'German',     'fr': 'French',     'es': 'Spanish',    'pt': 'Portuguese',
    'it': 'Italian',    'nl': 'Dutch',       'tr': 'Turkish',    'ja': 'Japanese',
    'zh': 'Chinese',    'ko': 'Korean',      'ru': 'Russian',    'ar': 'Arabic',
    'en': 'English',    'br': 'Portuguese',  # br used as lang code on ManageEngine
    # Nordic / Eastern European languages (e.g. /dk/ pages use lang="da")
    'da': 'Danish',     'sv': 'Swedish',     'no': 'Norwegian',  'nb': 'Norwegian',
    'fi': 'Finnish',    'el': 'Greek',       'pl': 'Polish',     'cs': 'Czech',
    'hu': 'Hungarian',  'ro': 'Romanian',    'sk': 'Slovak',     'bg': 'Bulgarian',
    'hr': 'Croatian',   'sl': 'Slovenian',   'lt': 'Lithuanian', 'lv': 'Latvian',
    'et': 'Estonian',
    # Asian / other languages
    'th': 'Thai',       'id': 'Indonesian',  'ms': 'Malay',      'vi': 'Vietnamese',
    'hi': 'Hindi',
}

# ISO 3166-1 alpha-2 country codes → primary page language for ManageEngine regional sites.
# Only country codes that differ from ISO 639-1 language codes need an entry here.
COUNTRY_LANG_MAP = {
    'at': 'German',     # Austria
    'au': 'English',    # Australia
    'be': 'Dutch',      # Belgium (nl-BE)
    'ca': 'English',    # Canada
    'ch': 'German',     # Switzerland
    'cl': 'Spanish',    # Chile
    'co': 'Spanish',    # Colombia
    'dk': 'Danish',     # Denmark
    'gb': 'English',    # United Kingdom
    'ie': 'English',    # Ireland
    'in': 'English',    # India
    'mx': 'Spanish',    # Mexico
    'my': 'English',    # Malaysia
    'ng': 'English',    # Nigeria
    'nz': 'English',    # New Zealand
    'pe': 'Spanish',    # Peru
    'ph': 'English',    # Philippines
    'se': 'Swedish',    # Sweden
    'sg': 'English',    # Singapore
    'za': 'English',    # South Africa
    'ae': 'Arabic',     # UAE
    'sa': 'Arabic',     # Saudi Arabia
}

# Combined list — order matters for CSV column output
EXPECTED_ELEMENTS = list(dict.fromkeys(
    ADAUDITPLUS_ELEMENTS + ADMANAGER_ELEMENTS + ADSELFSERVICE_ELEMENTS
    + EVENTLOG_ELEMENTS + LOG360_ELEMENTS + DATASECURITY_ELEMENTS
))


ELEMENT_ALIASES = {
    "active directory": [
        "active directory", "active-directory",
        "active directoryserveur",
    ],
    "file server": [
        "file server",
        "serveur de fichiers",           # French
        "dateiserver",                    # German
        "servidor de archivos",           # Spanish (ES, singular)
        "servidores de archivos",         # Spanish (LATAM, plural)
        "servidor de arquivos",           # Portuguese (BR)
        "servidor de ficheiros",          # Portuguese (PT)
        "server di file",                 # Italian
        "bestandsserver",                 # Dutch
        "dosya sunucusu",                 # Turkish
        "serwer plikow", "serwery plikow",  # Polish (Serwer plików / Serwery plików)
    ],
    "windows server": [
        "windows server", "windows-server",
        "serveur windows",                # French
        "servidor windows",               # Spanish (ES, singular) / Portuguese
        "servidores windows",             # Spanish (LATAM, plural)
        "windows sunucusu",               # Turkish
    ],
    "workstation": [
        "workstation",
        "station de travail", "poste de travail", "postes de travail",  # French
        "arbeitsstation", "arbeitsplatz",                                # German
        "estacion de trabajo",                                           # Spanish
        "estacao de trabalho",                                           # Portuguese
        "postazione di lavoro",                                          # Italian
        "werkstation",                                                   # Dutch
        "is istasyonu",                                                  # Turkish
        "stacja robocza", "stacje robocze",  # Polish (Stacja robocza / Stacje robocze)
    ],
    "compliance": [
        "compliance",
        "conformite", "conformité",      # French
        "einhaltung", "konformitat",     # German
        "cumplimiento",                  # Spanish
        "conformidade",                  # Portuguese
        "conformita", "conformità",      # Italian
        "naleving",                      # Dutch
        "uyumluluk",                     # Turkish
        "zgodnosc",                      # Polish (Zgodność / Zgodność z przepisami)
    ],
    "related products": [
        "related products",
        "produits connexes", "produits associes", "autres solutions",    # French
        "verwandte produkte", "zugehörige produkte", "zugehorige produkte",
        "ahnliche produkte", "andere losungen",    # German
        "productos relacionados", "productos afines",                    # Spanish
        "produtos relacionados", "produtos associados", "produtos populares",  # Portuguese (BR)
        "prodotti correlati",                                            # Italian
        "gerelateerde producten",                                        # Dutch
        "relaterede produkter",                                          # Danish
        "relaterade produkter",                                          # Swedish
        "relaterte produkter",                                           # Norwegian
        "ilgili urunler",                                                # Turkish
        # Non-Latin scripts — preserved by Unicode-aware normalize():
        "σχετικα προιοντα",    # Greek (Σχετικά προϊόντα, after NFD strip)
        "связанные продукты",  # Russian (Cyrillic)
        # Additional European languages
        "produkty powiazane",  # Polish (Produkty powiązane — actual ManageEngine form)
        "powiazane produkty",  # Polish alt word order
        "souvisejici produkty", # Czech
        "kapcsolodo termekek", # Hungarian
        "liittyvat tuotteet",  # Finnish
        "produse inrudite",    # Romanian
        "suvisiace produkty",  # Slovak
    ],
    "highlights": [
        "highlights",
        "points forts", "faits saillants",  # French
        "hoogtepunten",                      # Dutch
        "destacados",                        # Spanish (LATAM sub-pages, masculine)
        "destacadas",                        # Spanish (LATAM attribute pages, feminine)
        "funcionalidades destacadas",        # Spanish LATAM — div.highHea footer label
        "aspectos principales",             # Spanish (LATAM homepage, Cloud Security Plus)
        "destaques",                         # Portuguese
        "punti salienti",                    # Italian
        "one cikanlar",                      # Turkish
    ],
    "ad management": [
        "ad management", "ad-management",
        "ad-verwaltung", "ad verwaltung",    # German
        "gestion ad", "administration ad",   # French
        "administracion ad",                 # Spanish
        "gestion de ad",                     # Spanish LATAM — div.highHea label "Gestión de AD"
        "gerenciamento ad", "gerenciamento do ad",  # Portuguese
        "gestione ad",                       # Italian
        "ad beheer",                         # Dutch
        "ad yonetimi",                       # Turkish
    ],
    "active directory reports": [
        "active directory reports",
        "active directory berichte",              # German
        "rapports active directory",              # French
        "informes de active directory",           # Spanish
        "informes de ad",                         # Spanish LATAM — div.highHea label "Informes de AD"
        "relatorios do active directory",         # Portuguese
        "report di active directory",             # Italian
        "active directory rapporten",             # Dutch
        "active directory raporlari",             # Turkish
    ],
    "exchange management": [
        "exchange management",
        "gerenciamento do exchange", "gerenciamento exchange",  # Portuguese
        "gestion de ms exchange", "gestión de ms exchange",     # Spanish LATAM — div.highHea label
        # German pages use English tab label; avoiding "exchange verwaltung" as alias
        # because it false-matches "Exchange-Verwaltungstools" link text as a substring.
    ],
    "popular products": [
        "popular products",
        "beliebte produkte",     # German
        "produits populaires",   # French
        "productos populares",   # Spanish
        "produtos populares",    # Portuguese
        "prodotti popolari",     # Italian
        "populaire producten",   # Dutch
        "populer urunler",       # Turkish
    ],
    "password management": [
        "password management",
        "kennwortmanagement", "kennwort management",  # German
        "gestion des mots de passe",                  # French
        "gestion de contrasenas",                     # Spanish (ES)
        "administracion de contrasenas",              # Spanish (LATAM)
        "gerenciamento de senhas",                    # Portuguese
        "gestione delle password",                    # Italian
        "wachtwoordbeheer",                           # Dutch
        "parola yonetimi",                            # Turkish
    ],

    "adaptive mfa": [
        "adaptive mfa",
        "adaptiver mfa", "adaptives mfa",  # German
        "mfa adaptative",                   # French
        "mfa adaptativa",                   # Spanish (ES) / Portuguese
        "mfa adaptable",                    # Spanish (LATAM)
        "mfa adattivo",                     # Italian
        "adaptieve mfa",                    # Dutch
        "uyarlanabilir mfa",                # Turkish
    ],
    "quick links": [
        "quick links",
        "schnellzugriff", "schnelllinks",  # German
        "liens rapides",                    # French
        "enlaces rapidos",                  # Spanish (ES)
        "vinculos rapidos",                 # Spanish (LATAM)
        "links rapidos",                    # Portuguese
        "link rapidi",                      # Italian
        "snelkoppelingen",                  # Dutch
        "hizli baglantilar",                # Turkish
        "szybkie linki",                    # Polish
    ],
    "corporate sso": [
        "corporate sso",
        "enterprise sso",            # English (actual EN tab label)
        "firmen-sso", "firmen sso",  # German
        "sso entreprise",            # French
        "sso empresarial",           # Spanish (ES) / Portuguese
        "sso para empresas",         # Spanish (LATAM)
        "sso aziendale",             # Italian
        "zakelijke sso",             # Dutch
        "kurumsal sso",              # Turkish
    ],
    "log management": [
        "log management",
        "gerenciamento de logs",        # Portuguese (BR)
        "log-verwaltung",               # German
        "gestion des journaux",         # French
        "gestion de registros",         # Spanish (ES)
        "administracion de logs",       # Spanish (LATAM)
        "gestione log",                 # Italian
        "zarzadzanie dziennikami", "dzienniki zarzadzanie",  # Polish (both orders)
    ],
    "it compliance": [
        "it compliance",
        "conformidade de ti",           # Portuguese (BR)
        "it-konformitat",               # German
        "conformite it",                # French
        "cumplimiento de ti",           # Spanish
        "conformita it",                # Italian
        "zgodnosc it",                  # Polish
    ],
    "log analyzer": [
        "log analyzer", "log analysis", "log analyser",  # EN variants
        "analisador de logs",           # Portuguese (BR)
        "log-analysator", "log analysator",  # German
        "analyseur de journaux",        # French
        "analizador de registros",      # Spanish (ES)
        "analisis de logs",             # Spanish (LATAM)
        "analizzatore log",             # Italian
        "log analyse",                  # Danish/Dutch/Norwegian
        "analizator dziennikow",        # Polish
    ],
    # ADSelfService Plus 5th tab
    "self-service security": [
        "self-service security", "self service security",  # EN ("Self-service & security", & stripped)
        "autoservicio y seguridad", "autoservicio seguridad",  # Spanish (LATAM/ES)
        "autoatendimento e seguranca", "autoatendimento seguranca",  # Portuguese
        "selfservice et securite",     # French
        "selfservice und sicherheit",  # German
        "selfservice e sicurezza",     # Italian
        "selfservice en beveiliging",  # Dutch
        "selfservis ve guvenlik",      # Turkish
    ],
    # DataSecurity Plus tab aliases
    "solutions": [
        "solutions", "solution",
        "soluciones",                   # Spanish
        "solucoes", "soluções",         # Portuguese
        "losungen", "lösungen",         # German
        "soluzioni",                    # Italian
        "oplossingen",                  # Dutch
        "cozumler",                     # Turkish
    ],
    "regulatory compliance": [
        "regulatory compliance",
        "cumplimiento de la normativa", "cumplimiento normativo",   # Spanish (ES/LATAM)
        "conformidade regulatoria", "conformidade regulatória",     # Portuguese
        "conformite reglementaire", "conformité réglementaire",     # French
        "gesetzliche compliance", "regulatorische compliance",      # German
        "conformita normativa", "conformità normativa",             # Italian
        "regelgeving compliance",                                   # Dutch
        "duzenleyici uyumluluk",                                    # Turkish
    ],
    "resources": [
        "resources",
        "recursos",                     # Spanish / Portuguese
        "ressourcen",                   # German
        "ressources",                   # French
        "risorse",                      # Italian
        "hulpbronnen",                  # Dutch
        "kaynaklar",                    # Turkish
    ],
}


class FooterHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_footer = False
        self.footer_depth = 0
        self.footer_text_parts = []
        self.footer_links = []
        self.current_tag = None
        self.current_text = []
        self.stack = []

    def _is_footer_tag(self, tag, attrs):
        attrs_lower = {key.lower(): (value or "").lower() for key, value in attrs.items()}
        tag_lower = tag.lower()
        class_names = attrs_lower.get("class", "").split()
        role = attrs_lower.get("role", "")
        aria_label = attrs_lower.get("aria-label", "")
        data_testid = attrs_lower.get("data-testid", "")
        return (
            tag_lower in {"footer", "contentinfo"}
            or role in {"contentinfo", "footer"}
            or attrs_lower.get("id", "") == "footer"
            or "footer" in class_names
            or "footer" in aria_label
            or "footer" in data_testid
        )

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if self._is_footer_tag(tag, attrs):
            self.footer_depth += 1
            self.in_footer = True
        self.stack.append((tag.lower(), attrs))

    def handle_endtag(self, tag):
        tag_lower = tag.lower()
        if self._is_footer_tag(tag, {}):
            if self.footer_depth > 0:
                self.footer_depth -= 1
                self.in_footer = self.footer_depth > 0
        if self.stack:
            self.stack.pop()

    def handle_data(self, data):
        if self.in_footer:
            text = data.strip()
            if text:
                self.footer_text_parts.append(text)

    def handle_startendtag(self, tag, attrs):
        self.handle_starttag(tag, attrs)
        self.handle_endtag(tag)


class FooterAudit:
    def __init__(self, urls_path: str, output_path: str):
        self.urls_path = Path(urls_path)
        self.output_path = Path(output_path)

    @staticmethod
    def normalize(text: str) -> str:
        text = text.lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        # Use \w (Unicode word chars) instead of [a-z0-9] so non-Latin scripts
        # (Greek, Arabic, Cyrillic, etc.) are preserved and their aliases can match.
        text = re.sub(r"[^\w]+", " ", text).strip()
        return text

    @staticmethod
    def detect_language(url: str, html: str) -> str:
        """Return full language name.

        Priority (URL-first because ManageEngine often sets lang="en" on localised pages):
        1. URL path segment auto-resolution — most reliable for ManageEngine's structure:
             a. ISO 639-1 language codes (de, fr, nl, da…)  → LANG_CODE_MAP
             b. ISO 3166-1 country codes (dk, be, se…)       → COUNTRY_LANG_MAP
             c. Special region codes (latam, pt-br)           → URL_LANG_MAP
        2. <meta http-equiv="content-language">  → LANG_CODE_MAP lookup
        3. <html lang="..."> attribute  → LANG_CODE_MAP lookup
        4. Default: English
        """
        # 1. URL segment — always authoritative for ManageEngine regional pages.
        seg_m = re.search(r'manageengine\.com/([a-z]{2,6}(?:-[a-z]{2,6})?)', url, re.IGNORECASE)
        if seg_m:
            seg = seg_m.group(1).lower()
            if seg in LANG_CODE_MAP:
                return LANG_CODE_MAP[seg]
            if seg in COUNTRY_LANG_MAP:
                return COUNTRY_LANG_MAP[seg]
        for seg_str, name in URL_LANG_MAP:
            if seg_str in url:
                return name
        # 2. <meta http-equiv="content-language" content="...">
        m = re.search(r'content-language["\']?\s+content=["\']([^"\']+)["\']', html, re.IGNORECASE)
        if not m:
            m = re.search(
                r'<meta[^>]+content=["\']([a-z]{2,3}(?:-[A-Z]{2})?)["\'][^>]+'
                r'http-equiv=["\']content-language["\']', html, re.IGNORECASE)
        if m:
            code = m.group(1).split('-')[0].lower()
            if code in LANG_CODE_MAP:
                return LANG_CODE_MAP[code]
        # 3. <html lang="...">
        m = re.search(r'<html[^>]+lang=["\']([^"\']+)["\']', html, re.IGNORECASE)
        if m:
            code = m.group(1).split('-')[0].lower()
            if code in LANG_CODE_MAP:
                return LANG_CODE_MAP[code]
        return "English"

    @staticmethod
    def contains_keywords(text: str, keywords: list[str]) -> list[str]:
        normalized_text = FooterAudit.normalize(text)
        found = []
        for keyword in keywords:
            aliases = ELEMENT_ALIASES.get(keyword, [keyword])
            for alias in aliases:
                if FooterAudit.normalize(alias) in normalized_text:
                    found.append(keyword)
                    break
        return found

    _EMPTY_LHS = {"detected": False, "link_count": 0, "sections": [], "related_products": False}
    _EMPTY_RHS = {"detected": False, "link_count": 0, "sections": []}
    _EMPTY_CTA = {"detected": False, "pattern": "", "heading": "", "bullets": [], "cta_text": "", "form_present": False}

    def fetch_html(self, url: str) -> tuple[str, list[str], dict, dict, dict]:
        """Return (html, raw_tab_labels, lhs_data, rhs_data).

        raw_tab_labels: visible footer-nav tab texts from the rendered DOM (language-agnostic).
        lhs_data: dict with keys detected, link_count, sections, related_products.
        rhs_data: dict with keys detected, link_count, sections.
        Both sidebar dicts are empty / default when Playwright is unavailable (urllib fallback).
        """
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                # Note any redirect so callers can surface it; continue analysing the
                # destination page rather than erroring (e.g. locale KB pages that redirect
                # to the product page still show the product page's footer/CTA in a real browser).
                from urllib.parse import urlparse as _urlparse
                _req_path = _urlparse(url).path.rstrip('/')
                _fin_path = _urlparse(page.url).path.rstrip('/')
                _redirected_to = page.url if _req_path != _fin_path else None
                try:
                    page.wait_for_selector(
                        ".fea-nav, .footer-new, .pageTab, div.fea-nav-list, div.highHea",
                        timeout=15000
                    )
                except Exception:
                    pass
                # Detect ManageEngine "empty page" (HTTP 200 but custom error page).
                # Uses innerText on the live rendered DOM — more reliable than regex on serialized HTML.
                _is_empty = page.evaluate("""
                    () => {
                        const t = (document.body ? document.body.innerText : '').toLowerCase();
                        return t.includes('uh-oh') ||
                               t.includes('cette page est vide') ||
                               t.includes('page is empty') ||
                               t.includes('diese seite ist leer') ||
                               t.includes('esta página está vacía');
                    }
                """) or False
                if _is_empty:
                    browser.close()
                    raise URLError("page not found — ManageEngine empty page")
                # Remove ALL CSS-hidden product navs (display:none on self or any ancestor).
                page.evaluate("""
                    () => {
                        ['.fea-nav', '.footer-new', '.pageTab'].forEach(sel => {
                            document.querySelectorAll(sel).forEach(el => {
                                if (el.offsetParent === null) el.remove();
                            });
                        });
                    }
                """)
                # Wait for LHS related-products block to be injected by JS.
                # Product pages:  ul#lhsTree  with ul.relPrd / ul.releated-nav inside.
                # Help pages:     ul#lhsElement — no related-products block; falls through.
                try:
                    page.wait_for_function(
                        """() => {
                            const lhsTree = document.querySelector('ul#lhsTree');
                            if (lhsTree && lhsTree.querySelector('ul.relPrd, ul.releated-nav')) return true;
                            const sib = document.querySelector('ul#lhsRelPrd');
                            if (sib && sib.children.length > 0) return true;
                            if (document.querySelector('ul#lhsElement')) return true;
                            if (document.querySelector('ul#vMenu')) return true;
                            const helpPane = document.querySelector('div.help_left_pane');
                            if (helpPane && helpPane.querySelector('ul li a')) return true;
                            return false;
                        }""",
                        timeout=3000
                    )
                except Exception:
                    pass
                # Extract footer tab labels.
                raw_tabs = page.evaluate("""
                    () => {
                        const nav = document.querySelector('.fea-nav');
                        if (nav) {
                            const spans = Array.from(nav.querySelectorAll('.fea-nav-link span'))
                                .map(el => el.innerText.trim()).filter(Boolean);
                            if (spans.length) return spans;
                        }
                        const ptLi = Array.from(document.querySelectorAll('.pageTab li'))
                            .map(el => el.innerText.trim().replace(/\\s+/g, ' ')).filter(Boolean);
                        if (ptLi.length) return ptLi;
                        const ptA = Array.from(document.querySelectorAll('.pageTab li a'))
                            .map(el => el.innerText.trim()).filter(Boolean);
                        if (ptA.length) return ptA;
                        // Pattern 3: blog/article pages — div.fea-nav-list is ManageEngine's
                        // related-products section that appears instead of the full tabbed footer.
                        const feaNavList = document.querySelector('div.fea-nav-list');
                        if (feaNavList && feaNavList.querySelectorAll('a').length >= 3) {
                            return ['related products'];
                        }
                        // Pattern 4: LATAM/localized attribute & sub-pages — footer uses
                        // div.highHea > ul > li for tab labels (not .pageTab or .fea-nav).
                        const highHea = document.querySelector('div.highHea');
                        if (highHea) {
                            const tabs = Array.from(highHea.querySelectorAll('li'))
                                .map(el => el.innerText.trim()).filter(Boolean);
                            if (tabs.length) return tabs;
                        }
                        return [];
                    }
                """) or []
                # Sentinel: marks that Playwright successfully fetched the page but found
                # no footer tabs. Lets analyze_url() skip the infer_footer() fallback,
                # which can false-positive on body content of sem/lp landing pages.
                if not raw_tabs:
                    raw_tabs = ['__PLAYWRIGHT_NOFOOTER__']
                # Extract LHS sidebar navigation data.
                # Pattern A — product pages: ul#lhsTree (with span.ifTree section labels,
                #   ul.relPrd / ul.releated-nav related-product sub-lists).
                # Pattern B — help/doc pages: ul#lhsElement inside div.lhs (top-level
                #   <li> anchor texts are the section labels; no related-products block).
                lhs_data = page.evaluate("""
                    () => {
                        // Pattern A: product page LHS
                        const lhsTree = document.querySelector('ul#lhsTree');
                        if (lhsTree) {
                            // Skip if the LHS container is hidden — get-quote/landing pages
                            // keep ul#lhsTree in the DOM but hide the parent div.lhs-tree
                            // via display:none so the sidebar is not visible on those pages.
                            const _par = lhsTree.parentElement;
                            if (_par && window.getComputedStyle(_par).display === 'none') {
                                return {detected: false, link_count: 0, sections: [], related_products: false};
                            }
                            const navLinks = Array.from(lhsTree.querySelectorAll('a'))
                                .filter(a => !a.closest('ul.relPrd, ul.releated-nav'));
                            const sections = Array.from(lhsTree.querySelectorAll('span.ifTree'))
                                .map(el => el.innerText.trim()).filter(Boolean);
                            const relPrdInside = lhsTree.querySelector('ul.relPrd, ul.releated-nav');
                            const hasRelPrdInside = relPrdInside
                                ? relPrdInside.querySelectorAll('a').length > 0 : false;
                            let hasRelPrdSibling = false;
                            let sib = lhsTree.nextElementSibling;
                            while (sib) {
                                if ((sib.id === 'lhsRelPrd' || sib.classList.contains('relPrd')
                                        || sib.classList.contains('releated-nav'))
                                        && sib.querySelectorAll('a').length > 0) {
                                    hasRelPrdSibling = true; break;
                                }
                                sib = sib.nextElementSibling;
                            }
                            return {
                                detected: navLinks.length > 0,
                                link_count: navLinks.length,
                                sections: sections,
                                related_products: hasRelPrdInside || hasRelPrdSibling
                            };
                        }
                        // Helper: extract sections + relPro for help-page nav roots.
                        // Returns null if the element is zero-sized (hidden/off-screen template).
                        function helpPageLhs(root) {
                            const rect = root.getBoundingClientRect();
                            if (rect.width === 0 && rect.height === 0) return null;
                            const navLinks = Array.from(root.querySelectorAll('a'));
                            const sections = Array.from(root.children)
                                .filter(li => li.tagName === 'LI' && li.querySelector('a'))
                                .map(li => li.querySelector('a').innerText.trim())
                                .filter(Boolean);
                            let hasRelated = false;
                            let sib = root.nextElementSibling;
                            while (sib) {
                                if (sib.classList.contains('relPro')
                                        && sib.querySelectorAll('a').length > 0) {
                                    hasRelated = true; break;
                                }
                                sib = sib.nextElementSibling;
                            }
                            return { detected: navLinks.length > 0, link_count: navLinks.length,
                                     sections: sections, related_products: hasRelated };
                        }

                        // Pattern B: ADAudit Plus and similar help pages
                        const lhsEl = document.querySelector('ul#lhsElement');
                        if (lhsEl) { const r = helpPageLhs(lhsEl); if (r) return r; }

                        // Pattern C: ADManager Plus help pages (ul#vMenu must be rendered)
                        const vMenu = document.querySelector('ul#vMenu');
                        if (vMenu) { const r = helpPageLhs(vMenu); if (r) return r; }

                        // Pattern D: Cloud Log Management and similar help pages using
                        // div.help_left_pane (custom scrollbar sidebar — mCustomScrollbar).
                        const helpPane = document.querySelector('div.help_left_pane');
                        if (helpPane) {
                            const ul = helpPane.querySelector('ul');
                            if (ul) { const r = helpPageLhs(ul); if (r) return r; }
                        }

                        return {detected: false, link_count: 0, sections: [], related_products: false};
                    }
                """) or self._EMPTY_LHS
                # Extract RHS sidebar navigation data.
                # Pattern A — product pages: ul.rhsTree (id="rl") with p.rhsTreeHeader labels.
                # Pattern B — KB/help pages: ul.relPrd / ul.releated-nav appear in div.contentRgt
                #   (right column) rather than inside ul#lhsTree; same class names as LHS.
                rhs_data = page.evaluate("""
                    () => {
                        // Pattern A: product pages
                        const rhsTree = document.querySelector('ul.rhsTree');
                        if (rhsTree) {
                            const links = rhsTree.querySelectorAll('a').length;
                            const sections = Array.from(rhsTree.querySelectorAll('p.rhsTreeHeader'))
                                .map(el => el.innerText.trim()).filter(Boolean);
                            return { detected: links > 0, link_count: links, sections: sections };
                        }
                        // Pattern B: KB/help pages — find relPrd/releated-nav in the right column.
                        // Exclusions:
                        //   - inside lhsTree or its parent (LHS sidebar elements)
                        //   - inside any ancestor whose class contains "foot" (footer sections
                        //     on case-study pages use div.cusss-foot-com with ul.releated-nav)
                        const lhsTree = document.querySelector('ul#lhsTree');
                        const lhsParent = lhsTree ? lhsTree.parentElement : null;
                        function inFooter(el) {
                            let node = el.parentElement;
                            while (node && node !== document.body) {
                                if (node.tagName === 'FOOTER') return true;
                                for (const c of node.classList) {
                                    if (c.toLowerCase().includes('foot')) return true;
                                }
                                node = node.parentElement;
                            }
                            return false;
                        }
                        const rhsRelPrd = Array.from(
                            document.querySelectorAll('ul.relPrd, ul.releated-nav')
                        ).find(el => {
                            if (lhsTree && lhsTree.contains(el)) return false;
                            if (lhsParent && lhsParent.contains(el)) return false;
                            if (inFooter(el)) return false;
                            return true;
                        });
                        if (rhsRelPrd) {
                            const links = rhsRelPrd.querySelectorAll('a').length;
                            // Heading is in span.hea inside ul.relPrd
                            const heaSpan = rhsRelPrd.querySelector('span.hea');
                            const heading = heaSpan ? heaSpan.innerText.trim() : '';
                            return {
                                detected: links > 0,
                                link_count: links,
                                sections: heading ? [heading] : [],
                            };
                        }
                        return { detected: false, link_count: 0, sections: [] };
                    }
                """) or self._EMPTY_RHS
                # Wait briefly for JS-injected floating buttons before checking CTA
                # (page uses domcontentloaded so these may not be in DOM yet).
                try:
                    page.wait_for_selector("a.floading-btn", state="attached", timeout=2000)
                except Exception:
                    pass
                # Scroll to trigger scroll-activated CTA buttons, then wait for JS handlers.
                # On ebook/video landing pages floading-btn buttons are in the DOM but
                # permanently hidden — they stay invisible even after scroll.
                # On real CTA pages the scroll event makes them visible.
                page.evaluate("window.scrollTo(0, 600)")
                page.wait_for_timeout(1500)
                # Extract RHS Floating CTA — three patterns:
                # 1. div.rhs-content    — sticky form panel (heading, bullets, submit button)
                # 2. a.floading-btn     — fixed floating action buttons (demo / get-quote)
                # 3. div.sliding-buttons — sem/lp page sliding button panel + div.form-popup
                cta_data = page.evaluate("""
                    () => {
                        const empty = {detected: false, pattern: '', heading: '', bullets: [], cta_text: '', form_present: false};

                        // Pattern 1: sticky form panel (div.rhs-content with form structure).
                        // Requires div.rhs-form-head — ManageEngine's branded CTA heading div.
                        // Always present on real trial/demo panels; absent on ebook download panels
                        // (which use div.book-image + a hidden form that appears on button click).
                        const panel = document.querySelector('div.rhs-content');
                        if (panel) {
                            const form = panel.querySelector('form');
                            const btn  = panel.querySelector('input.ffw-submit, button[type="submit"]');
                            const formHead = panel.querySelector('div.rhs-form-head');
                            if (formHead) {
                                const h = panel.querySelector('div.rhs-form-head h2, div.rhs-form-head h3, h2, h3');
                                const bullets = Array.from(panel.querySelectorAll('div.rhs-form-head ul li'))
                                    .map(li => li.innerText.trim()).filter(Boolean);
                                return {
                                    detected: true,
                                    pattern: 'form-panel',
                                    heading: h ? h.innerText.trim() : '',
                                    bullets: bullets,
                                    cta_text: btn ? (btn.value || btn.innerText.trim()) : '',
                                    form_present: form !== null,
                                };
                            }
                        }

                        // Pattern 2: floating fixed buttons (a.floading-btn)
                        // Scroll-triggered buttons become visible after ~400px scroll.
                        // The page was already scrolled to 600px before this evaluate runs, so
                        // permanently-hidden buttons (ebook/landing pages) remain visibility:hidden
                        // and are filtered out; real CTA buttons are visible (or opacity>0) by now.
                        const floatBtns = Array.from(document.querySelectorAll('a.floading-btn'))
                            .filter(b => {
                                const s = window.getComputedStyle(b);
                                return s.visibility !== 'hidden' && s.display !== 'none' && parseFloat(s.opacity) > 0.1;
                            });
                        if (floatBtns.length > 0) {
                            // Derive label from href (language-neutral) so French/localized pages
                            // show consistent names instead of locale-specific data-value strings.
                            function ctaLabel(b) {
                                const pg = (b.getAttribute('href') || '').split('?')[0].split('/').pop().toLowerCase();
                                if (pg.includes('demo')) return 'Demo';
                                if (pg.includes('quote') || pg.includes('pricing')) return 'Get Quote';
                                if (pg.includes('trial') || pg.includes('free-trial')) return 'Free Trial';
                                if (pg.includes('download')) return 'Download';
                                if (pg.includes('callback')) return 'Callback';
                                return b.innerText.trim() || b.getAttribute('data-value') || pg;
                            }
                            const labels = floatBtns.map(ctaLabel).filter(Boolean);
                            return {
                                detected: true,
                                pattern: 'floating-buttons',
                                heading: '',
                                bullets: [],
                                cta_text: labels.join(' | '),
                                form_present: false,
                            };
                        }

                        // Pattern 3: sem/lp sliding button panel (div.sliding-buttons)
                        const slidingPanel = document.querySelector('div.sliding-buttons');
                        if (slidingPanel) {
                            const links = Array.from(slidingPanel.querySelectorAll('a'));
                            const labels = links.map(a => a.innerText.trim()).filter(Boolean);
                            const popup = document.querySelector('div.form-popup');
                            return {
                                detected: true,
                                pattern: 'sliding-buttons',
                                heading: '',
                                bullets: [],
                                cta_text: labels.join(' | '),
                                form_present: popup !== null,
                            };
                        }

                        // Pattern 4: fixed RHS pricing/quote panel (div#adRhsLnk)
                        const adRhs = document.querySelector('div#adRhsLnk');
                        if (adRhs) {
                            const s = window.getComputedStyle(adRhs);
                            if (s.display !== 'none' && s.visibility !== 'hidden') {
                                const labels = Array.from(adRhs.querySelectorAll('a'))
                                    .map(a => a.innerText.trim()).filter(Boolean);
                                return {
                                    detected: true,
                                    pattern: 'rhs-ad-panel',
                                    heading: '',
                                    bullets: [],
                                    cta_text: labels.join(' | '),
                                    form_present: false,
                                };
                            }
                        }

                        return empty;
                    }
                """) or self._EMPTY_CTA
                html = page.content()
                browser.close()
                return html, raw_tabs, lhs_data, rhs_data, cta_data
        except Exception:
            try:
                # urllib requires ASCII-only URLs — percent-encode any non-ASCII
                # characters (e.g. curly apostrophes in French/localized URLs).
                _p = urlparse(url)
                _ascii_url = urlunparse(_p._replace(path=quote(_p.path, safe='/:@!$&\'()*+,;=')))
                req = Request(_ascii_url, headers={"User-Agent": "Mozilla/5.0"})
                with urlopen(req, timeout=20) as resp:
                    return resp.read().decode("utf-8", errors="ignore"), [], self._EMPTY_LHS, self._EMPTY_RHS, self._EMPTY_CTA
            except (HTTPError, URLError) as exc:
                raise exc

    def infer_footer(self, html: str) -> tuple[bool, str, list[str], list[str]]:
        """Detect product navigation: HTML structure first, then tight text cluster, then footer tag."""
        # Strip inert content blocks before pattern matching.
        # HTML comments: ManageEngine comments-out the entire fea-nav section on sem/lp pages
        #   (e.g. <!--<section id="allFea">...<div class="fea-nav">...</div>-->).
        #   querySelectorAll never sees commented content, but regex does — causing false positives.
        # noscript: contains a static SEO copy of the nav that Playwright cannot remove.
        clean_html = re.sub(r'<!--.*?-->', '', html, flags=re.DOTALL)
        clean_html = re.sub(r'<noscript[^>]*>.*?</noscript>', '', clean_html, flags=re.DOTALL | re.IGNORECASE)
        clean_html = re.sub(r'<script[^>]*>.*?</script>', '', clean_html, flags=re.DOTALL | re.IGNORECASE)
        clean_html = re.sub(r'<style[^>]*>.*?</style>', '', clean_html, flags=re.DOTALL | re.IGNORECASE)

        text_only = re.sub(r'<[^>]+>', ' ', clean_html)
        text_only = re.sub(r'\s+', ' ', text_only).strip()

        found_footer = False
        footer_text = ""

        # Strategy 1: ManageEngine HTML structure markers (footer-new / pageTab).
        # Present on full product pages and sem/ pages; absent on sem/lp landing pages.
        nav_markers = [
            r'class=["\'][^"\']*footer-new[^"\']*["\']',   # ADAudit Plus
            r'class=["\'][^"\']*pageTab[^"\']*["\']',        # ADAudit Plus
            r'data-target=["\']#actDir["\']',                # ADAudit Plus
            r'data-target=["\']#rel-Prd["\']',               # ADAudit Plus
            r'class=["\'][^"\']*\bfea-nav\b[^"\']*["\']',   # ADManager Plus
            r'data-nav=["\']allFeaDiv',                       # ADManager Plus
        ]
        for marker in nav_markers:
            m = re.search(marker, clean_html, re.IGNORECASE)
            if m:
                start = max(0, m.start() - 200)
                end = min(len(clean_html), m.end() + 15000)
                block_text = re.sub(r'<[^>]+>', ' ', clean_html[start:end])
                block_text = re.sub(r'\s+', ' ', block_text).strip()
                if block_text:
                    found_footer = True
                    footer_text = block_text
                    break

        # Strategy 2: Tight text cluster — 4+ product elements within 200 chars,
        # searched only in the bottom 40% of the page. Product nav blocks always sit
        # near the bottom; early comparison sentences (e.g. competitor pages listing
        # all monitored systems) cannot trigger false positives.
        if not found_footer:
            TIGHT, STEP = 200, 100
            bottom_start = int(len(text_only) * 0.60)
            for pos in range(bottom_start, max(bottom_start + 1, len(text_only) - TIGHT), STEP):
                chunk = text_only[pos:pos + TIGHT]
                if len(set(self.contains_keywords(chunk, EXPECTED_ELEMENTS))) >= 4:
                    exp_start = max(0, pos - 200)
                    exp_end = min(len(text_only), pos + 5000)
                    found_footer = True
                    footer_text = text_only[exp_start:exp_end]
                    break

        # Strategy 3: HTML <footer> tag fallback — only when the tag contains at least
        # one product element. Copyright-only footers are not treated as detected.
        if not found_footer:
            parser = FooterHTMLParser()
            parser.feed(clean_html)
            tag_text = "\n".join(parser.footer_text_parts)
            if tag_text.strip() and self.contains_keywords(tag_text, EXPECTED_ELEMENTS):
                found_footer = True
                footer_text = tag_text

        if not found_footer or not footer_text.strip():
            return False, "", [], []

        found_elements = self.contains_keywords(footer_text, EXPECTED_ELEMENTS)
        matched = [elem for elem in EXPECTED_ELEMENTS if elem in found_elements]
        return True, footer_text, found_elements, matched

    def analyze_url(self, url: str) -> dict:
        try:
            html, raw_tabs, lhs_data, rhs_data, cta_data = self.fetch_html(url)
        except (HTTPError, URLError) as exc:
            seg_m = re.search(r'manageengine\.com/([a-z]{2,6}(?:-[a-z]{2,6})?)', url, re.IGNORECASE)
            lang_on_err = "English"
            if seg_m:
                seg = seg_m.group(1).lower()
                lang_on_err = LANG_CODE_MAP.get(seg) or COUNTRY_LANG_MAP.get(seg) or "English"
            if lang_on_err == "English":
                lang_on_err = next((n for s, n in URL_LANG_MAP if s in url), "English")
            return {
                "url": url,
                "language": lang_on_err,
                "footer_detected": False,
                "footer_status": "Missing Footer",
                "product": "",
                "detected_tabs": "",
                "active_directory": False,
                "file_server": False,
                "windows_server": False,
                "workstation": False,
                "compliance": False,
                "highlights": False,
                "ad_management": False,
                "active_directory_reports": False,
                "exchange_management": False,
                "popular_products": False,
                "password_management": False,
                "adaptive_mfa": False,
                "corporate_sso": False,
                "self_service_security": False,
                "log_management": False,
                "it_compliance": False,
                "log_analyzer": False,
                "quick_links": False,
                "ds_solutions": False,
                "ds_reg_compliance": False,
                "ds_resources": False,
                "ds_quick_links": False,
                "ds_related_products": False,
                "cs_highlights": False,
                "cs_related_products": False,
                "related_products": False,
                "lhs_detected": False,
                "lhs_link_count": 0,
                "lhs_sections": "",
                "lhs_related_products": False,
                "rhs_detected": False,
                "rhs_link_count": 0,
                "rhs_sections": "",
                "cta_detected": False,
                "cta_pattern": "",
                "cta_heading": "",
                "cta_bullets": "",
                "cta_text": "",
                "cta_form_present": False,
                "note": f"Fetch failed: {exc}",
            }

        language = self.detect_language(url, html)

        # Strip the Playwright sentinel before using raw_tabs.
        playwright_ran = '__PLAYWRIGHT_NOFOOTER__' in raw_tabs
        if playwright_ran:
            raw_tabs = []

        if raw_tabs:
            # Playwright found actual visible footer tabs — authoritative for any language.
            detected = True
            footer_text = ' '.join(raw_tabs)
            found_elements = self.contains_keywords(footer_text, EXPECTED_ELEMENTS)
        elif playwright_ran:
            # Playwright ran but found no tabs → footer is absent.
            # Do NOT fall back to infer_footer(): body-content keywords on sem/lp and
            # other landing pages can false-positive the keyword-cluster heuristic.
            detected = False
            footer_text = ''
            found_elements = []
        else:
            # urllib fallback (Playwright unavailable) — use HTML structure analysis.
            detected, footer_text, found_elements, _ = self.infer_footer(html)

        # URL-first product detection via PRODUCT_URL_MAP (extend the map to add new products).
        product = standard_elements = None
        for url_pat, pname, _, pstandard in PRODUCT_URL_MAP:
            if url_pat in url:
                product, standard_elements = pname, pstandard
                break

        if product is None:
            # Footer-element fallback for multi-tab products
            if any(e in found_elements for e in ADSELFSERVICE_CORE_ELEMENTS):
                product, standard_elements = "ADSelfService Plus", ADSELFSERVICE_CORE_ELEMENTS
            elif any(e in found_elements for e in ADMANAGER_ELEMENTS if e != "related products"):
                product, standard_elements = "ADManager Plus", ADMANAGER_CORE_ELEMENTS
            else:
                # Auto-derive product name from /REGION/SLUG/ — handles 2-char lang codes
                # (de, fr, br), 2+2 codes (pt-br), and multi-char region codes (latam, apac).
                slug_m = re.search(r'/(?:[a-z]{2,6}(?:-[a-z]{2,6})?/)([a-z][a-z0-9-]+)/', url)
                if slug_m:
                    slug = slug_m.group(1)
                    auto_name = ' '.join(w.capitalize() for w in slug.split('-'))
                    product, standard_elements = auto_name, SINGLE_TAB_ELEMENTS
                else:
                    product, standard_elements = "ADAudit Plus", ADAUDITPLUS_ELEMENTS

        # Core-check: require min(2,len) matching elements unless raw_tabs are available.
        # When Playwright extracted real DOM tabs the detection is already authoritative;
        # the alias-match count is irrelevant — some languages won't match any alias.
        if not raw_tabs:
            core_check = standard_elements
            min_required = min(2, len(core_check))
            if detected and sum(1 for e in core_check if e in found_elements) < min_required:
                detected = False
                found_elements = []

        if not detected:
            status = "Missing Footer"
        else:
            status = "Footer Detected"

        # Only report elements that belong to the detected product — prevents
        # link-text mentions of other products' terms from appearing as false positives.
        fe = set(found_elements) if detected else set()
        is_adaudit        = (product == "ADAudit Plus")
        is_admanager      = (product == "ADManager Plus")
        is_adselfservice  = (product == "ADSelfService Plus")
        is_eventlog       = (product == "EventLog Analyzer")
        is_datasecurity   = (product == "DataSecurity Plus")
        is_cloudsecurity  = (product == "Cloud Security Plus")
        return {
            "url": url,
            "language": language,
            "footer_detected": detected,
            "footer_status": status,
            "product": product,
            "detected_tabs": " | ".join(raw_tabs) if detected else "",
            # ADAudit Plus elements
            "active_directory": "active directory" in fe and is_adaudit,
            "file_server": "file server" in fe and is_adaudit,
            "windows_server": "windows server" in fe and is_adaudit,
            "workstation": "workstation" in fe and is_adaudit,
            "compliance": "compliance" in fe and is_adaudit,
            # ADManager Plus elements
            "highlights": "highlights" in fe and is_admanager,
            "ad_management": "ad management" in fe and is_admanager,
            "active_directory_reports": "active directory reports" in fe and is_admanager,
            "exchange_management": "exchange management" in fe and is_admanager,
            "popular_products": "popular products" in fe and is_admanager,
            # ADSelfService Plus elements
            "password_management": "password management" in fe and is_adselfservice,
            "adaptive_mfa": "adaptive mfa" in fe and is_adselfservice,
            "corporate_sso": "corporate sso" in fe and is_adselfservice,
            "self_service_security": "self-service security" in fe and is_adselfservice,
            # EventLog Analyzer elements
            "log_management": "log management" in fe and is_eventlog,
            "it_compliance": "it compliance" in fe and is_eventlog,
            "log_analyzer": "log analyzer" in fe and is_eventlog,
            "quick_links": "quick links" in fe and is_eventlog,
            # DataSecurity Plus elements
            "ds_solutions": "solutions" in fe and is_datasecurity,
            "ds_reg_compliance": "regulatory compliance" in fe and is_datasecurity,
            "ds_resources": "resources" in fe and is_datasecurity,
            "ds_quick_links": "quick links" in fe and is_datasecurity,
            "ds_related_products": "related products" in fe and is_datasecurity,
            # Cloud Security Plus elements
            "cs_highlights": "highlights" in fe and is_cloudsecurity,
            "cs_related_products": "related products" in fe and is_cloudsecurity,
            # related_products is shared — JS picks which section to show it in
            "related_products": "related products" in fe,
            # LHS sidebar navigation
            "lhs_detected": lhs_data.get("detected", False),
            "lhs_link_count": lhs_data.get("link_count", 0),
            "lhs_sections": " | ".join(lhs_data.get("sections", [])),
            "lhs_related_products": lhs_data.get("related_products", False),
            # RHS sidebar navigation
            "rhs_detected": rhs_data.get("detected", False),
            "rhs_link_count": rhs_data.get("link_count", 0),
            "rhs_sections": " | ".join(rhs_data.get("sections", [])),
            # RHS floating CTA (div.rhs-content form-panel or a.floading-btn buttons)
            "cta_detected": cta_data.get("detected", False),
            "cta_pattern": cta_data.get("pattern", ""),
            "cta_heading": cta_data.get("heading", ""),
            "cta_bullets": " | ".join(cta_data.get("bullets", [])),
            "cta_text": cta_data.get("cta_text", ""),
            "cta_form_present": cta_data.get("form_present", False),
            "note": "",
            "footer_excerpt": footer_text[:400] if detected else "",
        }

    def write_report(self, rows: list[dict], audit_type: str = "both"):
        lhs_cols = ["lhs_detected", "lhs_link_count", "lhs_sections", "lhs_related_products"]
        rhs_cols = ["rhs_detected", "rhs_link_count", "rhs_sections"]
        cta_cols = ["cta_detected", "cta_pattern", "cta_heading", "cta_bullets", "cta_text", "cta_form_present"]

        types = set(audit_type.split(","))
        do_footer = "footer" in types or audit_type == "both"
        do_lhs    = "lhs"    in types or audit_type == "both"
        do_rhs    = "rhs"    in types or audit_type == "both"
        do_cta    = "cta"    in types or audit_type == "both"

        # Build footer column groups (only if footer audit selected)
        footer_fieldnames = []
        if do_footer:
            base = ["url", "language", "footer_detected", "footer_status", "product", "detected_tabs"]
            product_cols = {
                "ADAudit Plus":       ["active_directory", "file_server", "windows_server", "workstation", "compliance"],
                "ADManager Plus":     ["highlights", "ad_management", "active_directory_reports", "exchange_management", "popular_products"],
                "ADSelfService Plus": ["password_management", "adaptive_mfa", "corporate_sso", "self_service_security"],
                "EventLog Analyzer":  ["log_management", "it_compliance", "log_analyzer", "quick_links"],
                "DataSecurity Plus":  ["ds_solutions", "ds_reg_compliance", "ds_resources", "ds_quick_links", "ds_related_products"],
                "Cloud Security Plus": ["cs_highlights", "cs_related_products"],
            }
            products_in_report = {r.get("product", "") for r in rows}
            tab_cols = []
            for product, cols in product_cols.items():
                if product in products_in_report:
                    tab_cols.extend(cols)
            footer_fieldnames = base + tab_cols + ["related_products", "note", "footer_excerpt"]
        else:
            footer_fieldnames = ["url", "language", "product"]

        sidebar_cols = (lhs_cols if do_lhs else []) + (rhs_cols if do_rhs else []) + (cta_cols if do_cta else [])
        # Merge: start with footer base, add sidebar cols (avoid duplicates)
        seen = set()
        fieldnames = []
        for col in footer_fieldnames + sidebar_cols:
            if col not in seen:
                seen.add(col)
                fieldnames.append(col)

        def _fmt(v):
            if v is True:  return "Available"
            if v is False: return "Not available"
            return v

        with self.output_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows({k: _fmt(v) for k, v in row.items()} for row in rows)

    @staticmethod
    def derive_canonical_url(url: str) -> str | None:
        """Replace a localised language segment with /products/ to get the English URL.

        Returns None if the URL is already English (no language segment found).
        Example: .com/fr/active-directory-audit/… → .com/products/active-directory-audit/…
        """
        m = re.match(r'(https://www\.manageengine\.com)/([a-z]{2,3})/(.*)', url, re.IGNORECASE)
        if m:
            return f"{m.group(1)}/products/{m.group(3)}"
        return None

    def write_comparison_report(self, rows: list[dict], output_path: "Path"):
        fieldnames = [
            "local_url", "english_url", "language", "product", "already_english",
            "local_load_error", "local_error_msg",
            "en_load_error", "en_error_msg",
            "local_lhs_detected", "local_lhs_links", "local_lhs_sections", "local_lhs_related",
            "en_lhs_detected",    "en_lhs_links",    "en_lhs_sections",    "en_lhs_related",
            "local_rhs_detected", "local_rhs_links", "local_rhs_sections",
            "en_rhs_detected",    "en_rhs_links",    "en_rhs_sections",
            "local_footer_detected", "local_footer_tabs",
            "en_footer_detected",    "en_footer_tabs",
            "local_cta_detected", "local_cta_pattern", "local_cta_text",
            "en_cta_detected",    "en_cta_pattern",    "en_cta_text",
        ]
        def _fmt(v):
            if v is True:  return "Available"
            if v is False: return "Not available"
            return v

        with output_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows({k: _fmt(v) for k, v in row.items()} for row in rows)

    def run(self, audit_type: str = "both"):
        urls = [line.strip() for line in self.urls_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            rows = list(executor.map(self.analyze_url, urls))
        self.write_report(rows, audit_type=audit_type)
        return rows


def _xl_color_map():
    """Return (hdr_fills, cell_fills, hdr_fonts) dicts keyed by section name.
    Uses 8-char ARGB strings (FF prefix = fully opaque) so PatternFill renders correctly."""
    from openpyxl.styles import PatternFill, Font
    # Same shade for headers and cells so each section column is uniformly coloured.
    _fills = {
        'lhs':     PatternFill(patternType='solid', fgColor='FFDBEAFE'),
        'rhs':     PatternFill(patternType='solid', fgColor='FFDCFCE7'),
        'footer':  PatternFill(patternType='solid', fgColor='FFFEF9C3'),
        'cta':     PatternFill(patternType='solid', fgColor='FFF3E4FF'),
        'general': PatternFill(patternType='solid', fgColor='FFE2E8F0'),
    }
    hdr_fills  = _fills
    cell_fills = {k: v for k, v in _fills.items()}
    cell_fills['general'] = None
    hdr_fonts = {
        'lhs':     Font(bold=True, color='FF1E40AF'),
        'rhs':     Font(bold=True, color='FF166534'),
        'footer':  Font(bold=True, color='FF854D0E'),
        'cta':     Font(bold=True, color='FF6B21A8'),
        'general': Font(bold=True, color='FF1E293B'),
    }
    return hdr_fills, cell_fills, hdr_fonts


def generate_excel_report(rows, fieldnames):
    """Generate a colored Excel workbook from the main audit report (rows already formatted as strings)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    LHS_COLS = {'lhs_detected', 'lhs_link_count', 'lhs_sections', 'lhs_related_products'}
    RHS_COLS = {'rhs_detected', 'rhs_link_count', 'rhs_sections'}
    CTA_COLS = {'cta_detected', 'cta_pattern', 'cta_heading', 'cta_bullets', 'cta_text', 'cta_form_present'}
    GENERAL  = {'url', 'language', 'product', 'footer_detected', 'footer_status',
                'related_products', 'note', 'footer_excerpt'}

    def cat(name):
        if name in LHS_COLS: return 'lhs'
        if name in RHS_COLS: return 'rhs'
        if name in CTA_COLS: return 'cta'
        if name in GENERAL:  return 'general'
        return 'footer'

    hdr_fills, cell_fills, hdr_fonts = _xl_color_map()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Footer Audit'
    ws.freeze_panes = 'A2'

    categories = [cat(f) for f in fieldnames]
    for ci, (fname, c) in enumerate(zip(fieldnames, categories), 1):
        cell = ws.cell(row=1, column=ci, value=fname)
        cell.fill = hdr_fills[c]
        cell.font = hdr_fonts[c]
        cell.alignment = Alignment(vertical='top')

    for ri, row in enumerate(rows, 2):
        for ci, (fname, c) in enumerate(zip(fieldnames, categories), 1):
            val = row.get(fname, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            if cell_fills[c]:
                cell.fill = cell_fills[c]
            cell.alignment = Alignment(vertical='top')

    sample = rows[:100]
    for ci, fname in enumerate(fieldnames, 1):
        col_vals = [len(fname)] + [len(str(r.get(fname, '') or '')) for r in sample]
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(col_vals) + 2, 50)

    return wb


def generate_comparison_excel(rows):
    """Generate a colored Excel workbook for the comparison report (booleans converted automatically)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    def _fmt(v):
        if v is True:  return 'Available'
        if v is False: return 'Not available'
        return v

    fieldnames = [
        'local_url', 'english_url', 'language', 'product', 'already_english',
        'local_load_error', 'local_error_msg', 'en_load_error', 'en_error_msg',
        'local_lhs_detected', 'local_lhs_links', 'local_lhs_sections', 'local_lhs_related',
        'en_lhs_detected',    'en_lhs_links',    'en_lhs_sections',    'en_lhs_related',
        'local_rhs_detected', 'local_rhs_links', 'local_rhs_sections',
        'en_rhs_detected',    'en_rhs_links',    'en_rhs_sections',
        'local_footer_detected', 'local_footer_tabs',
        'en_footer_detected',    'en_footer_tabs',
        'local_cta_detected', 'local_cta_pattern', 'local_cta_text',
        'en_cta_detected',    'en_cta_pattern',    'en_cta_text',
    ]

    def cat(name):
        if 'lhs'    in name: return 'lhs'
        if 'rhs'    in name: return 'rhs'
        if 'footer' in name: return 'footer'
        if 'cta'    in name: return 'cta'
        return 'general'

    hdr_fills, cell_fills, hdr_fonts = _xl_color_map()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comparison'
    ws.freeze_panes = 'A2'

    categories = [cat(f) for f in fieldnames]
    for ci, (fname, c) in enumerate(zip(fieldnames, categories), 1):
        cell = ws.cell(row=1, column=ci, value=fname)
        cell.fill = hdr_fills[c]
        cell.font = hdr_fonts[c]
        cell.alignment = Alignment(vertical='top')

    for ri, row in enumerate(rows, 2):
        for ci, (fname, c) in enumerate(zip(fieldnames, categories), 1):
            val = _fmt(row.get(fname, ''))
            cell = ws.cell(row=ri, column=ci, value=val)
            if cell_fills[c]:
                cell.fill = cell_fills[c]
            cell.alignment = Alignment(vertical='top')

    sample = rows[:100]
    for ci, fname in enumerate(fieldnames, 1):
        col_vals = [len(fname)] + [len(str(_fmt(r.get(fname, '')) or '')) for r in sample]
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(col_vals) + 2, 50)

    return wb


if __name__ == "__main__":
    input_path = sys.argv[1] if len(sys.argv) > 1 else "urls.txt"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "footer_report.csv"
    audit = FooterAudit(input_path, output_path)
    rows = audit.run()
    for row in rows:
        print(row)
